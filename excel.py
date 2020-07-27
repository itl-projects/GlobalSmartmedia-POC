import xlsxwriter
import openpyxl

def createSheet(name, data):
    workbook = xlsxwriter.Workbook(name)
    worksheet = workbook.add_worksheet()
    cols = 0
    for i in data.keys():
        worksheet.write(0, cols, i[0].upper() + i[1:].replace('_',' ').lower())
        cols += 1
    try:
        workbook.close()
        return True
    except:
        print("Couldn't create xlsx file")


def appendSheet(sheetName, data, row=None):
    wb = openpyxl.load_workbook(sheetName)
    sheetList = wb.get_sheet_names()
    sheet = wb.get_sheet_by_name(sheetList[0])
    temp = sheet.max_row
    cols = 1
    for i in data:
        cell = sheet.cell(row=temp+1, column=cols)
        cell.value = data[i]
        cols += 1
    temp += 1
    wb.save(sheetName)